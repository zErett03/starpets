"""
Бухгалтерия по заказам.

Считает выручку, закуп, комиссию и прибыль за период, плюс показывает деньги, застрявшие
в проблемных заказах.

ВАЖНОЕ ОГРАНИЧЕНИЕ. У заказа хранится ОДНА цена выкупа (exec_price_usd). Если предмет
перекупали — а мы это делаем при протухании и застрявших трейдах — предыдущая цена
затирается, и фактические затраты занижены. Точную картину дал бы журнал покупок, как
сделано в mmo-api; здесь его нет.
"""
from __future__ import annotations

from datetime import datetime, timedelta

import httpx
from fastapi import APIRouter
from sqlalchemy import select

from app.db import AsyncSessionLocal
from app.db.models import DeliveryStatus, Order

router = APIRouter()

CBR_URL = "https://www.cbr-xml-daily.ru/daily_json.js"

# Заказ считается принёсшим деньги только в этих статусах.
_DELIVERED = {DeliveryStatus.done, DeliveryStatus.finalized}
# Деньги покупателю уже вернули: closed — вручную оператором, failed — сорванная выдача.
# Выручки нет. Закуп при этом обычно НЕ теряется: если предмет не выдан и протух,
# StarPets возвращает его стоимость на баланс API. Реальный убыток — только комиссии,
# поэтому формула в таблице (-(H+I), без себестоимости) верна.
_REFUNDED = {DeliveryStatus.failed, DeliveryStatus.closed}
# Здесь деньги покупателя получены, но выдача не закрыта — риск возврата.
_AT_RISK = {DeliveryStatus.needs_attention}


def _period(since: str, until: str, days: int) -> tuple[datetime, datetime, str]:
    """Границы периода. Явные даты важнее относительных дней.

    Отбор идёт по дате ОПЛАТЫ, а не создания: бухгалтерия про деньги, а заказ мог быть
    заведён precheck-ом заранее.
    """
    def _parse(s: str, default: datetime) -> datetime:
        s = (s or "").strip()
        if not s:
            return default
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return default

    end = _parse(until, datetime.utcnow() + timedelta(days=1))
    if since.strip():
        start = _parse(since, datetime(1970, 1, 1))
    elif days > 0:
        start = datetime.utcnow() - timedelta(days=days)
    else:
        start = datetime(1970, 1, 1)
    label = f"{start:%d.%m.%Y} — {min(end, datetime.utcnow()):%d.%m.%Y}"
    return start, end, label


async def _usd_rate() -> tuple[float, str]:
    """Курс доллара ЦБ. Закуп у StarPets идёт в долларах, выручка в рублях."""
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(CBR_URL)
            resp.raise_for_status()
            v = (resp.json().get("Valute") or {}).get("USD") or {}
        rate = float(v.get("Value")) / float(v.get("Nominal") or 1)
        return rate, "cbr"
    except Exception:  # noqa: BLE001
        return 0.0, "unavailable"


@router.get("/accounting")
async def accounting(since: str = "", until: str = "", days: int = 30,
                     usd_rub: float = 0.0, commission: float = 0.028):
    """
    Итоги за период по дате оплаты. usd_rub=0 — курс ЦБ на сегодня.

    Комиссия по умолчанию 2.8% — платёжная система 2.7% плюс ggsel 0.1%, как в ручной
    таблице. Если фактические удержания окажутся иными, передай параметром.
    """
    rate, rate_src = (usd_rub, "manual") if usd_rub > 0 else await _usd_rate()
    start, end, period_label = _period(since, until, days)

    async with AsyncSessionLocal() as db:
        rows = (await db.execute(
            select(Order).where(Order.paid_at.isnot(None),
                                Order.paid_at >= start, Order.paid_at < end)
        )).scalars().all()

    delivered_n = at_risk_n = in_flight_n = unpaid_n = refunded_n = 0
    delivered_rev = delivered_cost_usd = 0.0
    at_risk_rev = at_risk_cost_usd = 0.0
    refunded_rev = refunded_cost_usd = 0.0
    spent_usd_total = 0.0

    for o in rows:
        sale = float(o.amount_rub or 0)
        cost_usd = float(o.exec_price_usd or 0)
        spent_usd_total += cost_usd
        if sale <= 0:
            # precheck заводит заказ на каждую попытку покупки; оплаченные имеют сумму.
            unpaid_n += 1
            continue
        if o.delivery_status in _DELIVERED:
            delivered_n += 1
            delivered_rev += sale
            delivered_cost_usd += cost_usd
        elif o.delivery_status in _REFUNDED:
            refunded_n += 1
            refunded_rev += sale
            refunded_cost_usd += cost_usd
        elif o.delivery_status in _AT_RISK:
            at_risk_n += 1
            at_risk_rev += sale
            at_risk_cost_usd += cost_usd
        else:
            in_flight_n += 1

    delivered_cost_rub = delivered_cost_usd * rate
    fee_rub = delivered_rev * commission
    profit = delivered_rev - fee_rub - delivered_cost_rub
    margin = (profit / delivered_rev) if delivered_rev else 0.0

    return {
        "period": period_label,
        "usd_rub": round(rate, 4),
        "usd_rub_source": rate_src,
        "commission": commission,
        "delivered": {
            "orders": delivered_n,
            "revenue_rub": round(delivered_rev, 2),
            "ggsel_fee_rub": round(fee_rub, 2),
            "purchase_cost_usd": round(delivered_cost_usd, 3),
            "purchase_cost_rub": round(delivered_cost_rub, 2),
            "profit_rub": round(profit, 2),
            "margin": round(margin, 4),
            "avg_check_rub": round(delivered_rev / delivered_n, 2) if delivered_n else 0,
            "avg_profit_rub": round(profit / delivered_n, 2) if delivered_n else 0,
        },
        "at_risk": {
            "orders": at_risk_n,
            "revenue_held_rub": round(at_risk_rev, 2),
            "purchase_cost_rub": round(at_risk_cost_usd * rate, 2),
            "note": ("оплачено покупателем, но выдача не закрыта: деньги под возвратом, "
                     "а предмет, возможно, уже куплен"),
        },
        "refunded": {
            "orders": refunded_n,
            "returned_rub": round(refunded_rev, 2),
            "purchase_cost_rub": round(refunded_cost_usd * rate, 2),
            "note": ("деньги возвращены покупателю, выручки нет. Стоимость закупа обычно "
                     "возвращается StarPets на баланс API, поэтому purchase_cost_rub — "
                     "не убыток, а справка. Убыток реален лишь если предмет успели выдать"),
        },
        "in_flight_orders": in_flight_n,
        "unpaid_precheck_orders": unpaid_n,
        "total_spent_usd_incl_failed": round(spent_usd_total, 3),
        "caveats": [
            "у заказа хранится ОДНА цена выкупа: перекупы (rebuy-fresh) затирают "
            "предыдущую, поэтому реальные затраты выше показанных",
            "курс берётся текущий и применяется ко всем заказам периода, хотя покупались "
            "они по курсам своих дней",
            "комиссия считается по ставке из параметра, фактические удержания ggsel "
            "могут отличаться — сверь с выпиской кабинета",
        ],
    }


@router.get("/accounting/xlsx")
async def accounting_xlsx(since: str = "", until: str = "", days: int = 0,
                          usd_rub: float = 0.0,
                          fee_ps: float = 0.027, fee_ggsel: float = 0.001):
    """
    Выгрузка в тот же формат, что и ручная таблица: три листа, формулы, параметры.

    Формулы сохранены живыми, а не посчитанными — чтобы можно было поменять комиссию или
    курс в «Параметрах» и увидеть пересчёт, как в исходном файле.

    days=0 — все заказы.
    """
    import io
    from fastapi.responses import StreamingResponse
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return {"error": "openpyxl не установлен — добавь его в requirements.txt"}

    rate, rate_src = (usd_rub, "manual") if usd_rub > 0 else await _usd_rate()
    start, end, period_label = _period(since, until, days)

    async with AsyncSessionLocal() as db:
        rows = (await db.execute(
            select(Order)
            .where(Order.paid_at.isnot(None), Order.paid_at >= start, Order.paid_at < end)
            .order_by(Order.paid_at.asc())
        )).scalars().all()

    # В таблице всего два учитываемых статуса. Остальные заказы попадают в файл, но в
    # суммы не входят: они ещё не закрыты, и записывать их в прибыль или убыток рано.
    def _status(o) -> str:
        if o.delivery_status in _DELIVERED:
            return "Доставлен"
        if o.delivery_status in _REFUNDED:      # failed + closed (возврат оператором)
            return "Возврат"
        return "В работе"

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    headers = ["№ заказа ggsel", "Товар", "Цена продажи P, ₽", "Курс USD/RUB",
               "Себест. выкупа, $", "Статус", "Себест., ₽", "Комиссия ПС, ₽",
               "Комиссия ggsel, ₽", "Net-выручка, ₽", "Прибыль, ₽", "Маржа %",
               "Оплачен"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    n = 0
    for o in rows:
        sale = float(o.amount_rub or 0)
        if sale <= 0:
            continue          # precheck-заготовки без оплаты в бухгалтерию не идут
        n += 1
        r = n + 1
        ws.append([
            o.ggsel_order_id, o.item_name, sale, round(rate, 4),
            float(o.exec_price_usd or 0), _status(o),
            f"=E{r}*D{r}",
            f"=C{r}*'Параметры'!$B$4",
            f"=C{r}*'Параметры'!$B$5",
            f"=C{r}-H{r}-I{r}",
            f'=IF($F{r}="Возврат",-(H{r}+I{r}),J{r}-G{r})',
            f"=IF(C{r}=0,0,K{r}/C{r})",
            o.paid_at.strftime("%d.%m.%Y %H:%M") if o.paid_at else "",
        ])

    last = n + 1
    s = wb.create_sheet("Сводка")
    s["A1"] = "Сводка (P&L)"
    s["A1"].font = Font(bold=True, size=13)
    s["A2"] = f"Период по дате оплаты: {period_label}"
    pairs = [
        (3, "Заказов всего", f"=COUNTA('Заказы'!$A$2:$A${last})"),
        (4, "  из них доставлено", f"=COUNTIF('Заказы'!$F$2:$F${last},\"Доставлен\")"),
        (5, "  из них возвраты", f"=COUNTIF('Заказы'!$F$2:$F${last},\"Возврат\")"),
        (6, "  из них в работе", f"=COUNTIF('Заказы'!$F$2:$F${last},\"В работе\")"),
        (8, "Валовая выручка (доставленные)",
            f"=SUMIFS('Заказы'!$C$2:$C${last},'Заказы'!$F$2:$F${last},\"Доставлен\")"),
        (9, "Комиссии всего (доставленные)",
            f"=SUMIFS('Заказы'!$H$2:$H${last},'Заказы'!$F$2:$F${last},\"Доставлен\")"
            f"+SUMIFS('Заказы'!$I$2:$I${last},'Заказы'!$F$2:$F${last},\"Доставлен\")"),
        (10, "Net-выручка (доставленные)",
             f"=SUMIFS('Заказы'!$J$2:$J${last},'Заказы'!$F$2:$F${last},\"Доставлен\")"),
        (11, "Себестоимость (доставленные)",
             f"=SUMIFS('Заказы'!$G$2:$G${last},'Заказы'!$F$2:$F${last},\"Доставлен\")"),
        (12, "Убыток от возвратов",
             f"=-SUMIFS('Заказы'!$K$2:$K${last},'Заказы'!$F$2:$F${last},\"Возврат\")"),
        (14, "ЧИСТАЯ ПРИБЫЛЬ",
             f"=SUMIFS('Заказы'!$K$2:$K${last},'Заказы'!$F$2:$F${last},\"Доставлен\")"
             f"+SUMIFS('Заказы'!$K$2:$K${last},'Заказы'!$F$2:$F${last},\"Возврат\")"),
        (15, "Средняя маржа (доставленные)",
             f"=IFERROR((SUMIFS('Заказы'!$J$2:$J${last},'Заказы'!$F$2:$F${last},\"Доставлен\")"
             f"-SUMIFS('Заказы'!$G$2:$G${last},'Заказы'!$F$2:$F${last},\"Доставлен\"))"
             f"/SUMIFS('Заказы'!$C$2:$C${last},'Заказы'!$F$2:$F${last},\"Доставлен\"),0)"),
    ]
    for row, label, formula in pairs:
        s.cell(row=row, column=1, value=label)
        s.cell(row=row, column=2, value=formula)
    s["A14"].font = Font(bold=True)
    s["B14"].font = Font(bold=True)
    s.column_dimensions["A"].width = 34
    s.column_dimensions["B"].width = 18

    p = wb.create_sheet("Параметры")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    p["A1"] = "Параметры расчёта"
    p["A1"].font = Font(bold=True, size=13)
    p["A2"] = "Жёлтые ячейки — редактируемые. Проценты хранятся как доли (0.027 = 2.7%)."
    params = [
        (4, "Комиссия платёжной системы", fee_ps, "Удерживается платёжкой с суммы заказа"),
        (5, "Комиссия ggsel", fee_ggsel, "Комиссия площадки ggsel"),
        (6, "Итого комиссия с продажи", "=B4+B5", "Сумма двух комиссий выше"),
        (7, "Курс USD/RUB", round(rate, 4), f"источник: {rate_src}"),
    ]
    for row, label, val, note in params:
        p.cell(row=row, column=1, value=label)
        c = p.cell(row=row, column=2, value=val)
        if isinstance(val, float):
            c.fill = yellow
        p.cell(row=row, column=3, value=note)
    notes = [
        (10, "Как считается прибыль"),
        (11, "Net-выручка = Цена продажи × (1 − итого комиссия). Комиссии при возврате НЕ возвращаются."),
        (12, "Себестоимость = цена выкупа StarPets ($) × курс USD/RUB."),
        (13, "Прибыль (доставлен) = Net-выручка − Себестоимость."),
        (14, "Прибыль (возврат) = −(комиссия ПС + комиссия ggsel)."),
        (16, "ОГРАНИЧЕНИЕ: у заказа хранится ОДНА цена выкупа. Перекупы (rebuy-fresh)"),
        (17, "затирают предыдущую, поэтому себестоимость таких заказов занижена."),
        (18, "Курс один на весь период, хотя заказы покупались по курсам своих дней."),
    ]
    for row, text in notes:
        p.cell(row=row, column=1, value=text)
    p["A10"].font = Font(bold=True)
    p["A16"].font = Font(bold=True)
    p.column_dimensions["A"].width = 78

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    for col in "CDEFGHIJKLM":
        ws.column_dimensions[col].width = 15
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = f"accounting_{datetime.utcnow():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@router.get("/accounting/push-sheets")
async def accounting_push_sheets(days: int = 14):
    """Отправить заказы в Google-таблицу вручную. По расписанию это делает воркер."""
    from app.workers.sheets_sync import push_orders
    return await push_orders(days=days)


@router.get("/accounting/orders")
async def accounting_orders(days: int = 30, only_problems: bool = False, limit: int = 200):
    """Построчная выгрузка — чтобы сверить итоги и найти, где деньги застряли."""
    since = datetime.utcnow() - timedelta(days=days)
    async with AsyncSessionLocal() as db:
        rows = (await db.execute(
            select(Order).where(Order.created_at >= since)
            .order_by(Order.id.desc()).limit(limit)
        )).scalars().all()

    out = []
    for o in rows:
        if only_problems and o.delivery_status not in _AT_RISK:
            continue
        if float(o.amount_rub or 0) <= 0 and not only_problems:
            continue
        out.append({
            "id": o.id,
            "ggsel_order_id": o.ggsel_order_id,
            "item": o.item_name,
            "status": o.delivery_status.value if o.delivery_status else None,
            "sale_rub": float(o.amount_rub) if o.amount_rub else None,
            "cost_usd": float(o.exec_price_usd) if o.exec_price_usd else None,
            "paid_at": o.paid_at.isoformat() if o.paid_at else None,
            "delivered_at": o.delivered_at.isoformat() if o.delivered_at else None,
            "error": (o.error_reason or "")[:120] or None,
        })
    return {"count": len(out), "orders": out}
